"""Codes that allow me to modify my spreadsheet(reworking) as well as looking up the new stock price"""
import openpyxl
import yfinance as yf   # downloading data from yahoo finance
import generalfunction

#theFile = openpyxl.load_workbook("stock portfolio.xlsx")
#allSheetNames = theFile.sheetnames
#currentSheet = 0


#for sheet in allSheetNames:
 #   if sheet == "stock portfolio":
  #      currentSheet = theFile[sheet]


#def find_specific_cell(str):
 #   for row in range(1, currentSheet.max_row + 1):
  #      for column in "ABCDE":
   #         cell_name = "{}{}".format(column, row)
    #        if currentSheet[cell_name].value == str:
     #           return cell_name
    #return 0

#def get_column_letter(specificCellLetter):
 #   if specificCellLetter == 0:
  #      return 0
   # letter = specificCellLetter[0:-1]
    #return letter

#def get_role_value(specificCellLetter):
 #   if specificCellLetter == 0:
  #      return 0
   # number = specificCellLetter[1:2]
    #return number

#def get_all_values_by_cell_letter(letter):
 #   for row in range(1, currentSheet.max_row + 1):
  #      for column in letter:
   #         cell_name = "{}{}".format(column, row)
    #        #print(cell_name)
     #       print("cell position {} has value {}".format(cell_name, currentSheet[cell_name].value))


def main():
    action = 1
    while action != 0:
        print("What do you want to do today\n")
        action = generalfunction.getnumber("\npress 1 to search current and historic stock price \npress 2 to use the calculator \n0 to quit ")
        print(action)
        #if action == 1:
            #Whattofind = input("What you want to find? ")
            #specificCellletter = (find_specific_cell(Whattofind))
            #value = get_role_value(specificCellletter)
            #newvalue = input("What the new value? ")
            #currentSheet['C' + str(value)] = float(newvalue)
            #theFile.save('stock portfolio.xlsx')
            #print("The stock " + Whattofind + " value has changed to " +newvalue)
        #elif action == 2:
            #Whattofind = input("What you want to find? ")
            #specificCellletter = (find_specific_cell(Whattofind))
            #value = get_role_value(specificCellletter)
            #price = currentSheet['C' + str(value)].value
            #print("\nThe value of the stock "+ Whattofind +" is "+ str(price) +"\n")
        if action == 1:
            Whattofind = input("Which stock price are you interested? ")
            whattofindprice = yf.Ticker(Whattofind)
            print("The price of " + Whattofind + " is\n" + str(whattofindprice.history(period="10d")))
            print("Power by yahoo finance")
            option = input("Would you want more data? \npress y for yes \npress n for no \n")
            if option == "y":
                period = generalfunction.getnumber("How many days of data do you want? ")
                print("The price of " + Whattofind + " is\n" + str(whattofindprice.history(period=(str(period) +"d"))))
        elif action ==2:
            calc = input("Type Calculation: \n")
            print("Answer: " + str(eval(calc)))





